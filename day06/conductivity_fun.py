import math
import openpyxl

def input_values():
    thickness = float(input("enter the thickness of sample in mm: "))
    diameter = float(input("enter the diameter of sample in mm: "))
    return(thickness, diameter)

def custom_factor(thickness, diameter):
    radius = diameter/2
    area = radius*radius*math.pi
    factor = (thickness/area)*10
    return(factor)
   
def conductivity_calculator(factor, Res):
    cond = factor/Res
    return(cond)

def excel_work(filename, factor):
    wb = openpyxl.load_workbook(filename)
    for sheet in wb.worksheets:
        temperature=273+int(sheet.title)
        min_value = None
        for row in sheet.iter_rows(min_row=1, min_col=2, max_col=2):
            for cell in row:
                if min_value is None or cell.value < min_value:
                    min_value = cell.value
                    Res = sheet.cell(row=cell.row, column=1).value
        conductivity = round(conductivity_calculator(factor, Res),6)
        print(f"For the temperature {temperature}K proton conductivity is {conductivity} S/cm.")
    return(conductivity)
    #returning value for doing tests with pytest